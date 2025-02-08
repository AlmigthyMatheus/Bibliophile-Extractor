import requests
from bs4 import BeautifulSoup
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from time import sleep
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Configure logging to display debug information
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


def get_session_with_retries(retries=3, backoff_factor=0.3):
    """
    Create a requests session with retry logic using exponential backoff.
    This helps to handle transient network errors gracefully.
    """
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=[500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def scrape_books():
    """
    Scrapes book data from "Books to Scrape" by iterating through all paginated pages.

    For each book, it extracts:
      - Title: from the 'title' attribute of the <a> tag inside the <h3> tag.
      - Price: from the <p> tag with class "price_color".
      - Rating: by reading the class from the <p> tag with class "star-rating"
                and converting it to star emojis.

    Returns:
        A list of dictionaries, where each dictionary contains:
        "Title", "Price", and "Rating" keys.
    """
    books = []
    page = 1
    base_url = "http://books.toscrape.com/catalogue/page-{}.html"
    session = get_session_with_retries()

    # Mapping from textual rating to star emojis
    star_mapping = {
        "One": "⭐",
        "Two": "⭐⭐",
        "Three": "⭐⭐⭐",
        "Four": "⭐⭐⭐⭐",
        "Five": "⭐⭐⭐⭐⭐"
    }

    while True:
        url = base_url.format(page)
        logging.info(f"Scraping page {page} from {url}")
        response = session.get(url)
        if response.status_code != 200:
            logging.warning("No more pages to scrape or encountered an error.")
            break

        soup = BeautifulSoup(response.text, "html.parser")
        articles = soup.find_all("article", class_="product_pod")
        if not articles:
            break

        for article in articles:
            # Extract title from the 'title' attribute
            title = article.h3.a["title"]
            # Extract price (e.g., "£51.77")
            price = article.find("p", class_="price_color").text.strip()
            # Extract rating by reading the second class from the star-rating <p> tag
            rating_classes = article.find("p", class_="star-rating").get("class", [])
            rating_text = rating_classes[1] if len(rating_classes) > 1 else "Not rated"
            # Convert textual rating to star emojis
            rating = star_mapping.get(rating_text, rating_text)
            books.append({
                "Title": title,
                "Price": price,
                "Rating": rating
            })

        page += 1
        sleep(0.5)  # Pause briefly to avoid overloading the server
    return books


def export_to_excel(data, output_file="books.xlsx"):
    """
    Exports the provided data to an Excel file with advanced formatting.

    Formatting includes:
      - Bold and centered headers.
      - All cells are center-aligned.
      - Automatic adjustment of column widths.
      - The "Rating" column is specially formatted with a gold font color.

    Args:
        data (list): A list of dictionaries containing the scraped book data.
        output_file (str): The name of the output Excel file.
    """
    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Export the DataFrame to Excel using openpyxl
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Books")
    writer.save()
    writer.close()  # Ensure the file is properly closed

    # Load the workbook to apply further formatting
    workbook = load_workbook(output_file)
    worksheet = workbook["Books"]

    # Define header style (bold) and default alignment (centered)
    header_font = Font(bold=True)
    default_alignment = Alignment(horizontal="center", vertical="center")

    # Format header row and identify the "Rating" column
    rating_col_letter = None
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = default_alignment
        if cell.value and cell.value.lower() == "rating":
            rating_col_letter = cell.column_letter

    # Adjust column widths based on content length and apply alignment
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
            cell.alignment = default_alignment
            # If this is the "Rating" column, apply a gold font color
            if rating_col_letter and cell.column_letter == rating_col_letter:
                cell.font = Font(color="FFD700", bold=(cell.row == 1))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(output_file)
    logging.info(f"Data successfully exported to {output_file}")


if __name__ == "__main__":
    logging.info("Starting Books to Scrape Web Scraping Example")
    books_data = scrape_books()
    if not books_data:
        logging.error("No data scraped.")
    else:
        for book in books_data:
            logging.debug(book)
        export_to_excel(books_data, output_file="books.xlsx")
